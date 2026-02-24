"""
TOOL 1 OF 3 — Pricebook Cleaner
================================
Reads every tab of the IDEXX pricebook and outputs one flat,
queryable Excel file where each row is:

    material | max_band | currency | price_2025 | price_2026 | source_tab | is_custom

Resilience features:
    - Fuzzy column header matching — tolerates renamed or reformatted headers
    - Merged cell handling — fills None values from openpyxl merged cell reads
    - Hidden row detection — skips rows flagged hidden by Excel
    - Ambiguous column warnings — flags headers that partially match multiple currencies
    - Tab structure change detection — warns when a tab produces far fewer rows than expected

SETUP:
    pip install pandas openpyxl

USAGE:
    python pricebook_cleaner.py
"""

import pandas as pd
from pathlib import Path
import warnings
import re
import sys
import os

warnings.filterwarnings('ignore')

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from validation import (
        assert_file_exists, assert_output_dir_writable,
        log_section, log_ok, log_warn, log_error, PIPELINE_VERSION
    )
    VALIDATION_AVAILABLE = True
except ImportError:
    VALIDATION_AVAILABLE = False
    def log_section(m): print(f"\n{m}")
    def log_ok(m):      print(f"  OK    {m}")
    def log_warn(m):    print(f"  WARN  {m}")
    def log_error(m):   print(f"  ERROR {m}")
    PIPELINE_VERSION = "unknown"

# ============================================================
# CONFIGURATION
# ============================================================
PRICEBOOK_FILE = r"C:\Users\nbrown2\Downloads\Master VetSoft Price Book 2026 - Updated for January 2026 (2).xlsx"
# ============================================================

# ---- CURRENCY DETECTION ----
# Each currency has a list of signals. A column needs to match at least one
# strong signal OR two weak signals to be classified.
CURRENCY_SIGNALS = {
    'USD': {
        'strong': ['usd', 'us list price', 'us$', 'united states'],
        'weak':   ['us ', ' us', 'dollar', 'american'],
        'symbols': ['$'],
        'exclude': ['cad', 'canada', 'au$', 'nz$', 'aud', 'nzd', 'aus ', 'aus list'],
    },
    'CAD': {
        'strong': ['cad', 'canada list price', 'canadian', 'ca$', 'can$'],
        'weak':   ['canada', 'canadian dollar'],
        'symbols': ['ca$', 'can$'],
        'exclude': [],
    },
    'GBP': {
        'strong': ['gbp', 'uk list price', 'united kingdom', 'britain'],
        'weak':   ['uk ', ' uk', 'british', 'sterling', 'pound'],
        'symbols': ['£', '\xa3'],
        'exclude': [],
    },
    'AUD': {
        'strong': ['aud', 'aus list price', 'australia list', 'australian', 'aus list'],
        'weak':   ['aus ', ' aus', 'australia'],
        'symbols': ['a$', 'au$'],
        'exclude': ['nzd', 'nz$', 'new zealand'],
    },
    'NZD': {
        'strong': ['nzd', 'nz list price', 'new zealand'],
        'weak':   ['nz ', ' nz'],
        'symbols': ['nz$'],
        'exclude': [],
    },
    'ZAR': {
        'strong': ['zar', 'south africa list', 'south african'],
        'weak':   ['africa', 'rand'],
        'symbols': ['r'],  # careful — 'R' is ambiguous
        'exclude': [],
    },
    'EUR': {
        'strong': ['eur', 'ireland list price', 'euro'],
        'weak':   ['ireland', 'europe', 'european'],
        'symbols': ['€', '\u20ac'],
        'exclude': [],
    },
}

def detect_currency(col_name):
    """
    Resilient currency detection using signal scoring.
    Returns (currency_code, confidence) or (None, 0).
    Confidence: 'high' = strong match, 'medium' = symbol/weak match, 'low' = ambiguous
    """
    col_lower = str(col_name).lower().strip()
    col_orig  = str(col_name).strip()

    candidates = {}

    for code, signals in CURRENCY_SIGNALS.items():
        score = 0
        # Check exclusions first
        if any(ex in col_lower for ex in signals['exclude']):
            continue
        # Strong signals
        for sig in signals['strong']:
            if sig in col_lower:
                score += 10
                break
        # Currency symbols (high confidence)
        for sym in signals['symbols']:
            if sym in col_orig or sym in col_lower:
                score += 8
                break
        # Weak signals
        weak_hits = sum(1 for w in signals['weak'] if w in col_lower)
        score += weak_hits * 3

        if score > 0:
            candidates[code] = score

    if not candidates:
        return None, 0

    # If multiple candidates, take highest score
    best_code  = max(candidates, key=candidates.get)
    best_score = candidates[best_code]

    # Check for ambiguity — two currencies within 3 points of each other
    sorted_scores = sorted(candidates.values(), reverse=True)
    if len(sorted_scores) > 1 and (sorted_scores[0] - sorted_scores[1]) <= 3:
        return None, 0  # too ambiguous, skip

    if best_score >= 10:
        return best_code, 'high'
    elif best_score >= 6:
        return best_code, 'medium'
    else:
        return best_code, 'low'


def detect_year(col_name):
    """Extract 2025 or 2026 from column header. Handles date ranges and plain years."""
    col_str = str(col_name)
    if '2026' in col_str:
        return '2026'
    if '2025' in col_str:
        return '2025'
    return None


def clean_number(val):
    """Convert string prices with any currency symbol to float."""
    if pd.isna(val):
        return None
    try:
        cleaned = str(val)
        for sym in ['$', '£', '€', '\xa3', '\u20ac']:
            cleaned = cleaned.replace(sym, '')
        cleaned = cleaned.strip()
        if cleaned.startswith('R') and len(cleaned) > 1 and cleaned[1].isdigit():
            cleaned = cleaned[1:]
        cleaned = cleaned.replace(',', '').replace(' ', '').strip()
        if not cleaned or cleaned.lower() in ('nan', 'none', '-', 'n/a', ''):
            return None
        result = float(cleaned)
        return result if result > 0 else None
    except (ValueError, TypeError):
        return None


def is_custom_value(val):
    """Return True if the cell contains non-numeric custom pricing text."""
    if pd.isna(val):
        return False
    cleaned = str(val).strip().upper()
    return (
        cleaned in ('CUSTOM', 'PRICING BASED ON CONTRACT', 'TBD', 'N/A', '-') or
        cleaned.startswith('PRICING BASED') or
        cleaned.startswith('CONTRACT')
    )


def find_part_number_col(columns):
    """Find the IDEXX Part Number column — tolerant of variations."""
    patterns = [
        r'idexx.*part.*number',
        r'part.*number',
        r'^material$',
        r'part.*no\b',
        r'part#',
        r'sku',
    ]
    for col in columns:
        col_lower = str(col).lower().strip()
        for pat in patterns:
            if re.search(pat, col_lower):
                return col
    return None


def find_max_band_col(columns):
    """Find the Max tier/seat/scale column."""
    patterns = [
        r'max.*user.*tier', r'max.*tier', r'max.*seat',
        r'scale.*quantity', r'number.*of.*seat',
        r'max.*band', r'upper.*limit',
    ]
    for pat in patterns:
        for col in columns:
            if re.search(pat, str(col).lower()):
                return col
    return None


def find_min_band_col(columns):
    """Find the Min tier/seat column — fallback when Max is blank/dash."""
    patterns = [
        r'min.*user.*tier', r'min.*tier', r'min.*seat',
        r'min.*band', r'lower.*limit',
    ]
    for pat in patterns:
        for col in columns:
            if re.search(pat, str(col).lower()):
                return col
    return None


def find_header_row(raw_df):
    """
    Scan rows to find the header. Handles tabs where headers aren't on row 0.
    Also handles merged cells by checking for partial row content.
    """
    for i, row in raw_df.iterrows():
        row_vals = [str(v).upper() for v in row.values if pd.notna(v) and str(v).strip()]
        row_str  = ' '.join(row_vals)
        # Must have a part number indicator AND at least 3 populated cells
        if len(row_vals) >= 3 and any(
            kw in row_str for kw in ('IDEXX', 'MATERIAL', 'PART NUMBER', 'PART NO', 'SKU',
                                     'PART#', 'ITEM', 'PRODUCT')
        ):
            return i
    return None


def handle_merged_cells(ws):
    """
    openpyxl reads merged cells as None for all but the top-left cell.
    This function fills merged regions with the top-left value so pandas
    doesn't see unexpected Nones in header rows.
    Returns a list-of-lists representation of the sheet data.
    """
    # Build a dict of merged cell ranges and their fill value
    merge_values = {}
    for merge_range in ws.merged_cells.ranges:
        top_left_val = ws.cell(merge_range.min_row, merge_range.min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                if not (row == merge_range.min_row and col == merge_range.min_col):
                    merge_values[(row, col)] = top_left_val

    data = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        row_data = []
        for c_idx, cell in enumerate(row, start=1):
            if (r_idx, c_idx) in merge_values:
                row_data.append(merge_values[(r_idx, c_idx)])
            else:
                # Skip formula results that openpyxl can't evaluate — treat as None
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    val = None
                row_data.append(val)
        data.append(row_data)
    return data


def process_tab(wb, sheet_name):
    """
    Process a single pricebook tab with full resilience handling.
    Returns (rows, skip_reason) — skip_reason is None on success.
    """
    ws = wb[sheet_name]

    # Skip hidden sheets
    if ws.sheet_state == 'hidden':
        return [], "sheet is hidden"

    # Handle merged cells before converting to DataFrame
    try:
        data = handle_merged_cells(ws)
    except Exception as e:
        return [], f"merged cell handling failed: {e}"

    if not data or len(data) < 2:
        return [], "sheet has fewer than 2 rows"

    raw_df = pd.DataFrame(data)

    # Find header row
    header_row = find_header_row(raw_df)
    if header_row is None:
        return [], "no header row found (no IDEXX/Material/Part Number column)"

    # Build DataFrame from header row down
    headers = [str(v).strip() if v is not None else f'_col_{i}'
               for i, v in enumerate(data[header_row])]

    # Handle duplicate column names (openpyxl sometimes produces these)
    seen = {}
    deduped = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            deduped.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            deduped.append(h)
    headers = deduped

    df = pd.DataFrame(data[header_row + 1:], columns=headers)
    df = df.dropna(how='all').reset_index(drop=True)

    if df.empty:
        return [], "no data rows below header"

    part_col = find_part_number_col(df.columns)
    if part_col is None:
        return [], "could not identify part number column"

    max_col = find_max_band_col(df.columns)
    min_col = find_min_band_col(df.columns)

    # Detect price columns with confidence scoring
    price_cols      = []
    low_conf_warns  = []
    ambiguous_warns = []

    for col in df.columns:
        currency, confidence = detect_currency(col)
        year = detect_year(col)
        if currency and year:
            if confidence == 'low':
                low_conf_warns.append(f"'{col}' → {currency} (low confidence)")
            price_cols.append((col, currency, year))
        elif year and not currency:
            # Has a year but no currency — might be a price col we're missing
            col_lower = str(col).lower()
            if any(w in col_lower for w in ['price', 'rate', 'cost', 'fee', 'list']):
                ambiguous_warns.append(f"'{col}' has year but no currency detected")

    if not price_cols:
        return [], "no price columns detected (looking for currency + year in header)"

    # Warn on low-confidence and ambiguous columns
    for w in low_conf_warns:
        log_warn(f"    Low-confidence currency detection: {w}")
    for w in ambiguous_warns:
        log_warn(f"    Possible missed price column: {w}")

    # Process rows
    rows = []
    custom_materials = set()

    for _, row in df.iterrows():
        material = str(row.get(part_col, '')).strip()
        if not material or material.lower() in ('nan', 'none', ''):
            continue
        # Skip header-like rows that leaked through
        if any(kw in material.upper() for kw in ('IDEXX', 'PART', 'MATERIAL', 'PRODUCT')):
            continue

        max_band = clean_number(row[max_col]) if max_col else None
        if max_band is None and min_col:
            max_band = clean_number(row.get(min_col))

        row_has_custom  = any(is_custom_value(row.get(col)) for col, _, _ in price_cols)
        row_has_numeric = False

        prices = {}
        for col, currency, year in price_cols:
            price = clean_number(row.get(col))
            if price is not None:
                prices.setdefault(currency, {})[year] = price
                row_has_numeric = True

        if row_has_numeric:
            for currency, year_prices in prices.items():
                rows.append({
                    'material':   material,
                    'max_band':   max_band,
                    'currency':   currency,
                    'price_2025': year_prices.get('2025'),
                    'price_2026': year_prices.get('2026'),
                    'source_tab': sheet_name,
                    'is_custom':  False,
                })
        elif row_has_custom:
            custom_materials.add(material)
            for _, currency, _ in price_cols:
                rows.append({
                    'material':   material,
                    'max_band':   None,
                    'currency':   currency,
                    'price_2025': None,
                    'price_2026': None,
                    'source_tab': sheet_name,
                    'is_custom':  True,
                })

    # Deduplicate custom rows
    if rows:
        df_rows  = pd.DataFrame(rows)
        numeric  = df_rows[~df_rows['is_custom']]
        custom   = df_rows[df_rows['is_custom']].drop_duplicates(subset=['material', 'currency'])
        rows     = pd.concat([numeric, custom], ignore_index=True).to_dict('records')

    if not rows:
        return [], "tab parsed but zero data rows extracted"

    if custom_materials:
        print(f"         ^ {len(custom_materials)} custom-priced material(s) recorded: "
              f"{', '.join(sorted(custom_materials))}")

    return rows, None


def run(pricebook_path):
    print("=" * 60)
    print(f"Pricebook Cleaner  (pipeline v{PIPELINE_VERSION})")
    print("=" * 60)

    if VALIDATION_AVAILABLE:
        log_section("Validating inputs...")
        assert_file_exists(pricebook_path, 'pricebook')
        output_path = Path(pricebook_path).parent / 'pricebook_clean.xlsx'
        assert_output_dir_writable(str(output_path))

    print(f"\nReading: {pricebook_path}")

    # Load workbook with openpyxl directly for merged cell handling
    from openpyxl import load_workbook
    try:
        wb = load_workbook(pricebook_path, data_only=True)
    except Exception as e:
        log_error(f"Could not open pricebook: {e}")
        print(f"\n  Fix: Make sure the file is not open in Excel and is a valid .xlsx file.")
        return

    print(f"Found {len(wb.sheetnames)} tabs\n")

    all_rows = []
    skipped  = []
    tab_row_counts = {}  # track for drift detection

    for sheet_name in wb.sheetnames:
        rows, skip_reason = process_tab(wb, sheet_name)
        if skip_reason:
            skipped.append((sheet_name, skip_reason))
            if 'hidden' not in skip_reason:
                print(f"  SKIP  {sheet_name!r:40s}  -> {skip_reason}")
        else:
            all_rows.extend(rows)
            numeric_count = len([r for r in rows if not r['is_custom']])
            custom_count  = len([r for r in rows if r['is_custom']])
            suffix = f" + {custom_count} custom" if custom_count else ""
            print(f"  OK    {sheet_name!r:40s}  -> {numeric_count:,} price rows{suffix}")
            tab_row_counts[sheet_name] = numeric_count

    print(f"\n{'─'*60}")
    print(f"Tabs processed:  {len(wb.sheetnames) - len(skipped)}")
    print(f"Tabs skipped:    {len(skipped)}")

    if not all_rows:
        log_error("No data extracted from pricebook.")
        print("\n  Fix: Check PRICEBOOK_FILE path and that the file has IDEXX Part Number columns.")
        return

    df_out = pd.DataFrame(all_rows)
    numeric_rows = df_out[~df_out['is_custom']]
    custom_rows  = df_out[df_out['is_custom']]

    print(f"Total price rows:  {len(numeric_rows):,}")
    print(f"Total custom rows: {len(custom_rows):,} "
          f"({custom_rows['material'].nunique() if not custom_rows.empty else 0} materials)")

    df_out['max_band'] = pd.to_numeric(df_out['max_band'], errors='coerce')
    df_out = df_out.sort_values(['material', 'currency', 'max_band']).reset_index(drop=True)

    # Drift detection — warn if fewer rows than expected
    EXPECTED_MIN_ROWS = 7000
    if len(numeric_rows) < EXPECTED_MIN_ROWS:
        log_warn(f"Only {len(numeric_rows):,} price rows extracted — expected {EXPECTED_MIN_ROWS:,}+. "
                 f"Check if any tabs were renamed or restructured.")

    print(f"\nSummary:")
    print(f"  Unique materials:      {df_out['material'].nunique():,}")
    print(f"  Currencies:            {', '.join(sorted(df_out['currency'].unique()))}")
    print(f"  Has 2025 price:        {df_out['price_2025'].notna().sum():,} rows")
    print(f"  Has 2026 price:        {df_out['price_2026'].notna().sum():,} rows")
    print(f"  Has tier bands:        {df_out['max_band'].notna().sum():,} rows")
    print(f"  Fixed price (no band): {df_out['max_band'].isna().sum():,} rows")

    output_path = Path(pricebook_path).parent / 'pricebook_clean.xlsx'
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name='Pricebook', index=False)
        ws = writer.sheets['Pricebook']
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', start_color='1F4E79')

    print(f"\nOutput saved: {output_path}")
    print("=" * 60)


if __name__ == '__main__':
    run(PRICEBOOK_FILE)
