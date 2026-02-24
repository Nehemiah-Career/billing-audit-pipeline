"""
TOOL 3 OF 3 — Audit Engine
===========================
Joins pricebook_clean.xlsx + sap_clean.xlsx and flags every billing row.

FLAGS:
    CORRECT_2026            - billed at current 2026 price
    PRICE_UNCHANGED         - 2025 and 2026 price are the same, matches both
    OLD_PRICE_2025          - billed at last year's price
    NO_MATCH                - doesn't match either year
    CUSTOM_PRICING          - material exists in pricebook but is manually priced
    NOT_IN_PRICEBOOK        - material not found in pricebook at all
    NO_TIER_BAND_FOUND      - material found but no band covers this quantity
    ZERO_QTY_FLAT_PRICE     - qty=0, price doesn't match either year
    CREDIT                  - negative net value
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
warnings.filterwarnings('ignore')
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from validation import (
        validate_pricebook_clean, validate_sap_clean, validate_audit_result,
        write_run_log, assert_output_dir_writable, run_with_error_handling,
        log_section, log_ok, log_warn, log_error, PIPELINE_VERSION
    )
    VALIDATION_AVAILABLE = True
except ImportError:
    VALIDATION_AVAILABLE = False
    def log_section(m): print(f"\n{m}")
    def log_ok(m): print(f"  OK    {m}")
    def log_warn(m): print(f"  WARN  {m}")
    def log_error(m): print(f"  ERROR {m}")
    PIPELINE_VERSION = "unknown"


# ============================================================
# CONFIGURATION
# ============================================================
PRICEBOOK_CLEAN = r"C:\Users\nbrown2\Downloads\pricebook_clean.xlsx"
SAP_CLEAN       = r"C:\Users\nbrown2\OneDrive - IDEXX\sap_clean.xlsx"
OUTPUT_FILE     = r"C:\Users\nbrown2\Downloads\billing_audit.xlsx"
# ============================================================

TOLERANCE = 0.01

FLAG_STYLES = {
    'CORRECT_2026':         {'bg': 'D9EAD3', 'font': '3D6B35'},  # soft green
    'PRICE_UNCHANGED':      {'bg': 'DDEBF7', 'font': '2E5F8A'},  # soft blue
    'OLD_PRICE_2025':       {'bg': 'FFF2CC', 'font': '7D6608'},  # soft yellow
    'NO_MATCH':             {'bg': 'F4CCCC', 'font': '6B1A1A'},  # muted rose
    'CUSTOM_PRICING':       {'bg': 'FCE5CD', 'font': '8A4A00'},  # soft amber
    'NOT_IN_PRICEBOOK':     {'bg': 'EAD1DC', 'font': '6B1A3A'},  # soft mauve
    'NO_TIER_BAND_FOUND':   {'bg': 'EAD1DC', 'font': '6B1A3A'},  # soft mauve
    'ZERO_QTY_FLAT_PRICE':  {'bg': 'E8F4FD', 'font': '2E5F8A'},  # pale blue
    'CREDIT':               {'bg': 'EFEFEF', 'font': '595959'},  # light grey
    'NO_PRICEBOOK_CURRENCY':{'bg': 'E8E8E8', 'font': '595959'},  # neutral grey - pricebook gap not billing error
    'BILLED_AT_ZERO':       {'bg': 'FFF2CC', 'font': '7D6608'},  # soft yellow - charged $0
}

HEADER_BG   = '1F4E79'
HEADER_FONT = 'FFFFFF'

OUTPUT_COLUMN_ORDER = [
    'status', 'address', 'sales_org', 'created_on', 'sold_to', 'ship_to',
    'order_number', 'material', 'customer_name', 'description',
    'quantity', 'net_value', 'currency', 'cost_group',
    'tier_max_band', 'price_2025', 'price_2026', 'source_tab',
    'audit_flag', 'variance_vs_2026',
]


def find_tier_price(quantity, pb_slice):
    banded   = pb_slice[pb_slice['max_band'].notna()].copy()
    unbanded = pb_slice[pb_slice['max_band'].isna()].copy()
    if not banded.empty:
        banded_sorted = banded.sort_values('max_band')
        matches = banded_sorted[banded_sorted['max_band'] >= quantity]
        row = matches.iloc[0] if not matches.empty else banded_sorted.iloc[-1]
        return row['max_band'], row['price_2025'], row['price_2026']
    if not unbanded.empty:
        row = unbanded.iloc[0]
        return None, row['price_2025'], row['price_2026']
    return None, None, None


def classify(net_value, price_2025, price_2026, quantity=None):
    if net_value < 0:
        return 'CREDIT'
    if net_value == 0:
        return 'BILLED_AT_ZERO'
    if price_2025 is None and price_2026 is None:
        return 'NO_TIER_BAND_FOUND'
    match_2025 = price_2025 is not None and abs(net_value - price_2025) <= TOLERANCE
    match_2026 = price_2026 is not None and abs(net_value - price_2026) <= TOLERANCE
    # Also check quantity * unit price (for per-unit billing)
    if not match_2025 and not match_2026 and quantity and quantity > 0:
        if price_2026 is not None and abs(net_value - (price_2026 * quantity)) <= TOLERANCE:
            return 'CORRECT_2026'
        if price_2025 is not None and abs(net_value - (price_2025 * quantity)) <= TOLERANCE:
            match_2025, match_2026 = True, False
            # re-check 2026 after qty multiply already done above
        if price_2026 is not None and price_2025 is not None and            abs(net_value - (price_2026 * quantity)) <= TOLERANCE:
            return 'CORRECT_2026'
    if match_2025 and match_2026:
        return 'PRICE_UNCHANGED'
    if match_2026:
        return 'CORRECT_2026'
    if match_2025:
        return 'OLD_PRICE_2025'
    return 'NO_MATCH'


def run():
    print("=" * 65)
    print(f"Billing Audit Engine  (pipeline v{PIPELINE_VERSION})")
    print("=" * 65)

    # Stage handoff validation
    if VALIDATION_AVAILABLE:
        assert_output_dir_writable(OUTPUT_FILE)
        validate_pricebook_clean(PRICEBOOK_CLEAN)
        validate_sap_clean(SAP_CLEAN)
    else:
        log_warn("validation.py not found — running without validation checks")
        log_warn("Copy validation.py to the same folder as audit_engine.py")

    pb = pd.read_excel(PRICEBOOK_CLEAN, sheet_name='Pricebook', dtype={'material': str})
    pb['material'] = pb['material'].str.strip()
    pb['currency'] = pb['currency'].str.strip().str.upper()
    pb['max_band'] = pd.to_numeric(pb['max_band'], errors='coerce')
    # Ensure is_custom column exists even on older pricebook_clean files
    if 'is_custom' not in pb.columns:
        pb['is_custom'] = False
    pb['is_custom'] = pb['is_custom'].fillna(False)

    numeric_pb = pb[~pb['is_custom']]
    custom_pb  = pb[pb['is_custom']]
    custom_materials = set(custom_pb['material'].unique())

    print(f"\nPricebook: {len(numeric_pb):,} price rows | {pb['material'].nunique()} materials")
    print(f"  Custom-priced materials: {len(custom_materials)}: "
          f"{', '.join(sorted(custom_materials)) if custom_materials else 'none'}")
    print(f"  Currencies: {sorted(pb['currency'].unique())}")

    sap = pd.read_excel(SAP_CLEAN, sheet_name='SAP', dtype={'material': str})
    sap['material'] = sap['material'].str.strip()
    sap['currency'] = sap['currency'].str.strip().str.upper()
    print(f"\nSAP: {len(sap):,} rows | {sap['material'].nunique()} materials")
    print(f"  Currencies: {sorted(sap['currency'].unique())}")

    pb_mats  = set(pb['material'].unique())
    sap_mats = set(sap['material'].unique())
    truly_missing = sap_mats - pb_mats
    print(f"\n  SAP materials found in pricebook: {len(sap_mats - truly_missing)} / {len(sap_mats)}")
    print(f"  SAP materials NOT in pricebook:   {sorted(truly_missing)}")

    print(f"\nRunning audit...")
    results = []
    no_match_samples = []

    for _, row in sap.iterrows():
        material  = str(row['material']).strip()
        currency  = str(row['currency']).strip().upper()
        quantity  = row['quantity']
        net_value = row['net_value']

        result = row.to_dict()

        # Check if custom-priced (material in pricebook but all prices are Custom)
        if material in custom_materials:
            result.update({
                'tier_max_band':    None,
                'price_2025':       None,
                'price_2026':       None,
                'variance_vs_2026': None,
                'source_tab':       custom_pb[custom_pb['material'] == material]['source_tab'].iloc[0],
                'audit_flag':       'CUSTOM_PRICING',
            })
            results.append(result)
            continue

        # Filter numeric pricebook rows
        pb_slice = numeric_pb[
            (numeric_pb['material'] == material) &
            (numeric_pb['currency'] == currency)
        ]

        if pb_slice.empty:
            pb_mat = numeric_pb[numeric_pb['material'] == material]
            if pb_mat.empty:
                flag = 'NOT_IN_PRICEBOOK'
                src  = None
            else:
                # Material exists but no price for this currency = pricebook gap, not billing error
                flag = 'NO_PRICEBOOK_CURRENCY'
                src  = pb_mat['source_tab'].iloc[0]
                if len(no_match_samples) < 5:
                    avail = sorted(pb_mat['currency'].unique())
                    no_match_samples.append(
                        f"    material={material} | SAP currency={currency} | PB has={avail}"
                    )
            result.update({
                'tier_max_band':    None,
                'price_2025':       None,
                'price_2026':       None,
                'variance_vs_2026': None,
                'source_tab':       src,
                'audit_flag':       flag,
            })
        else:
            if quantity == 0:
                row_pb     = pb_slice.iloc[0]
                price_2025 = row_pb['price_2025']
                price_2026 = row_pb['price_2026']
                tier_max   = None
                flag       = classify(net_value, price_2025, price_2026)
                if flag == 'NO_MATCH':
                    flag = 'ZERO_QTY_FLAT_PRICE'
                elif flag == 'BILLED_AT_ZERO':
                    flag = 'BILLED_AT_ZERO'
            else:
                tier_max, price_2025, price_2026 = find_tier_price(quantity, pb_slice)
                flag = classify(net_value, price_2025, price_2026, quantity)

            variance = (
                round(net_value - price_2026, 2)
                if price_2026 is not None else None
            )
            result.update({
                'tier_max_band':    tier_max,
                'price_2025':       price_2025,
                'price_2026':       price_2026,
                'variance_vs_2026': variance,
                'source_tab':       pb_slice['source_tab'].iloc[0],
                'audit_flag':       flag,
            })

        results.append(result)

    if no_match_samples:
        print(f"\n  DEBUG — NO_MATCH currency mismatches (first 5):")
        for s in no_match_samples:
            print(s)

    audit = pd.DataFrame(results)
    ordered   = [c for c in OUTPUT_COLUMN_ORDER if c in audit.columns]
    remaining = [c for c in audit.columns if c not in ordered]
    audit     = audit[ordered + remaining]

    flag_counts   = audit['audit_flag'].value_counts()
    correct_flags = {'CORRECT_2026', 'PRICE_UNCHANGED'}
    correct_count = audit[audit['audit_flag'].isin(correct_flags)].shape[0]

    print(f"\n  {'Flag':<30} {'Rows':>6}   {'Total Billed':>14}")
    print(f"  {'─'*55}")
    for flag, count in flag_counts.items():
        total = audit[audit['audit_flag'] == flag]['net_value'].sum()
        print(f"  {flag:<30} {count:>6,}   ${total:>13,.2f}")
    print(f"  {'─'*55}")
    print(f"  {'TOTAL':<30} {len(audit):>6,}   ${audit['net_value'].sum():>13,.2f}")
    print(f"\n  Clean matches (CORRECT_2026 + PRICE_UNCHANGED): {correct_count} (expecting ~557)")

    review_flags = {'OLD_PRICE_2025', 'NO_MATCH', 'NOT_IN_PRICEBOOK',
                    'NO_TIER_BAND_FOUND', 'CUSTOM_PRICING', 'BILLED_AT_ZERO', 'NO_PRICEBOOK_CURRENCY'}
    correct_set  = {'CORRECT_2026', 'PRICE_UNCHANGED', 'ZERO_QTY_FLAT_PRICE', 'CREDIT'}
    needs_review = audit[audit['audit_flag'].isin(review_flags)].copy()
    correct      = audit[audit['audit_flag'].isin(correct_set)].copy()

    if VALIDATION_AVAILABLE:
        validate_audit_result(audit, len(sap))
        write_run_log(
            Path(OUTPUT_FILE).parent,
            {
                'sap_rows':    len(sap),
                'correct':     flag_counts.get('CORRECT_2026', 0) + flag_counts.get('PRICE_UNCHANGED', 0),
                'no_match':    flag_counts.get('NO_MATCH', 0),
                'custom':      flag_counts.get('CUSTOM_PRICING', 0),
                'old_price':   flag_counts.get('OLD_PRICE_2025', 0),
                'total_billed': audit['net_value'].sum(),
            }
        )

    print(f"\nWriting output: {OUTPUT_FILE}")
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        needs_review.to_excel(writer, sheet_name='Needs_Review', index=False)
        correct.to_excel(writer,      sheet_name='Correct',      index=False)
        audit.to_excel(writer,        sheet_name='Full_Data',    index=False)

        summary_rows = []
        for flag, count in flag_counts.items():
            subset = audit[audit['audit_flag'] == flag]
            summary_rows.append({
                'Audit Flag':   flag,
                'Row Count':    count,
                '% of Rows':    f"{count/len(audit)*100:.1f}%",
                'Total Billed': round(subset['net_value'].sum(), 2),
                'Avg Billed':   round(subset['net_value'].mean(), 2),
            })
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name='Summary', index=False)

    wb = load_workbook(OUTPUT_FILE)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row <= 1:
            continue
        for cell in ws[1]:
            cell.font      = Font(bold=True, color=HEADER_FONT, name='Arial')
            cell.fill      = PatternFill('solid', start_color=HEADER_BG)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        flag_col_idx = None
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value == 'audit_flag':
                flag_col_idx = col[0].column
                break

        if flag_col_idx and sheet_name != 'Summary':
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                flag_val = row[flag_col_idx - 1].value
                style    = FLAG_STYLES.get(flag_val)
                if style:
                    for cell in row:
                        cell.fill = PatternFill('solid', start_color=style['bg'])
                        cell.font = Font(color=style['font'], name='Arial')
                    row[flag_col_idx - 1].font = Font(
                        bold=True, color=style['font'], name='Arial'
                    )

        if sheet_name == 'Summary':
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                flag_val = row[0].value
                style    = FLAG_STYLES.get(flag_val)
                if style:
                    for cell in row:
                        cell.fill = PatternFill('solid', start_color=style['bg'])
                        cell.font = Font(color=style['font'], name='Arial')

        for col in ws.columns:
            max_len = max(
                len(str(c.value)) if c.value is not None else 0 for c in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

    wb.move_sheet('Needs_Review', offset=-wb.index(wb['Needs_Review']))
    wb.active = wb['Needs_Review']
    wb.save(OUTPUT_FILE)
    print(f"\n  Done. Open: {OUTPUT_FILE}")
    print("=" * 65)


if __name__ == '__main__':
    run()
