"""
TOOL 2 OF 3 — SAP Data Cleaner
================================
Reads your SAP billing export and outputs a clean, standardized
Excel file ready to join against pricebook_clean.xlsx.

Resilience features:
    - Extra header rows: scans up to row 30 for the real header
    - Shifted columns: re-scans all rows if initial header scan fails
    - Missing columns: reports ALL missing columns at once with fix suggestions
    - Partial column matches: warns when a column name is close but not exact
    - Subtotal row detection: drops SAP-injected subtotal/total rows
    - European number format: handles both 1,234.56 and 1.234,56
    - Blank rows mid-export: dropped automatically
    - Multi-sheet exports: scans all sheets for the one with billing data
    - Unknown currency codes: warns before audit runs

SETUP:
    pip install pandas openpyxl

USAGE:
    python sap_cleaner.py
"""

import pandas as pd
from pathlib import Path
import warnings
import re
import sys
import os

warnings.filterwarnings('ignore')

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from validation import (
        assert_file_exists, assert_output_dir_writable,
        log_section, log_ok, log_warn, log_error, PIPELINE_VERSION
    )
    VALIDATION_AVAILABLE = True
except ImportError:
    VALIDATION_AVAILABLE = False
    def log_section(m): print(f"\n{m}")
    def log_ok(m):      print(f"  OK    {m}")
    def log_warn(m):    print(f"  WARN  {m}")
    def log_error(m):   print(f"  ERROR {m}")
    PIPELINE_VERSION = "unknown"

# ============================================================
# CONFIGURATION
# ============================================================
SAP_FILE = r"C:\Users\nbrown2\OneDrive - IDEXX\sap export jan.xlsx"
# ============================================================

# Required columns with broad pattern lists — survives SAP config changes
COL_PATTERNS = {
    'material':  ['MATERIAL', 'PART NUMBER', 'PART NO', 'PART#', 'SKU', 'ITEM NO',
                  'ITEM NUMBER', 'PRODUCT NO', 'MATL'],
    'quantity':  ['ORDER QUANT', 'ORDER QTY', 'ORDERQUAN', 'QUANTITY', 'QTY',
                  'UNITS', 'ORDERED QTY', 'BILLED QTY', 'BILL QTY'],
    'net_value': ['NET VALUE', 'NETVALUE', 'NET VAL', 'NET AMT', 'NET AMOUNT',
                  'BILLED AMT', 'BILLED AMOUNT', 'AMOUNT', 'REVENUE', 'BILLING AMT'],
    'currency':  ['CURR', 'CURRENCY', 'CCY', 'CUR'],
}

# Context columns — carried through if found, not required
CONTEXT_PATTERNS = {
    'sales_org':     ['SORG', 'S ORG', 'SALES ORG', 'SALESORG'],
    'created_on':    ['CREATED ON', 'CREATEDON', 'CREATE DATE', 'CREATION DATE'],
    'order_number':  ['ORDER#', 'ORDER #', 'ORDER NUMBER', 'ORDERNO', 'ORDER NO',
                      'DOCUMENT', 'DOC NO', 'BILLING DOC'],
    'ship_to':       ['SHIP-TO', 'SHIP TO', 'SHIPTO'],
    'customer_name': ['NAME 1', 'NAME1', 'CUSTOMER NAME', 'CUSTOMER', 'CLIENT NAME'],
    'address':       ['ADDRESS', 'ADDR'],
    'status':        ['ST.', 'STATUS', 'STS'],
    'sold_to':       ['SOLD TO', 'SOLDTO', 'SOLD-TO', 'PAYER'],
    'description':   ['DESCRIPTION', 'SAP DESC', 'DESC', 'PRODUCT DESC',
                      'MATERIAL DESC', 'TEXT', 'ITEM DESC'],
    'cost_group':    ['CGP', 'COST GROUP', 'COSTGROUP', 'COST GRP'],
}

SUBTOTAL_INDICATORS = [
    'subtotal', 'sub total', 'total', 'grand total',
    'sum', 'count', '**', '***', 'page total',
]

CURRENCY_NORMALIZATIONS = {
    'GBP': 'GBP', 'USD': 'USD', 'CAD': 'CAD',
    'AUD': 'AUD', 'NZD': 'NZD', 'ZAR': 'ZAR', 'EUR': 'EUR',
    'US$': 'USD', 'CA$': 'CAD', 'AU$': 'AUD', 'NZ$': 'NZD',
    'EURO': 'EUR', 'RAND': 'ZAR',
}


def clean_number(val, european_format=False):
    if pd.isna(val):
        return None
    cleaned = str(val).strip()
    is_negative = cleaned.startswith('(') and cleaned.endswith(')')
    if is_negative:
        cleaned = cleaned[1:-1]
    for sym in ['$', '£', '€', '\xa3', '\u20ac']:
        cleaned = cleaned.replace(sym, '')
    cleaned = cleaned.strip()
    if cleaned.startswith('R') and len(cleaned) > 1 and cleaned[1].isdigit():
        cleaned = cleaned[1:]
    if not european_format:
        comma_pos  = cleaned.rfind(',')
        period_pos = cleaned.rfind('.')
        if comma_pos > period_pos and comma_pos > 0:
            european_format = True
    if european_format:
        cleaned = cleaned.replace('.', '').replace(',', '.')
    else:
        cleaned = cleaned.replace(',', '')
    cleaned = cleaned.replace(' ', '').strip()
    if not cleaned or cleaned.lower() in ('nan', 'none', '-', 'n/a', ''):
        return None
    try:
        result = float(cleaned)
        return -result if is_negative else result
    except (ValueError, TypeError):
        return None


def find_header_row(raw_df, sheet_name=''):
    """Scan up to row 30 for a row containing material AND currency indicators."""
    for i, row in raw_df.head(30).iterrows():
        row_upper = ' '.join(str(v).upper() for v in row.values if pd.notna(v))
        has_material = any(k in row_upper for k in
                          ['MATERIAL', 'PART NUMBER', 'PART NO', 'SKU', 'MATL'])
        has_value    = any(k in row_upper for k in
                          ['NET VALUE', 'NETVALUE', 'CURR', 'CURRENCY', 'AMOUNT'])
        if has_material and has_value:
            if i > 0:
                log_warn(f"Header found at row {i+1} — {i} junk row(s) skipped above it")
            return i
    return None


def find_data_sheet(xl):
    """Find the sheet that contains billing data in a multi-sheet export."""
    for sheet_name in xl.sheet_names:
        try:
            raw = xl.parse(sheet_name, header=None, nrows=30)
            for _, row in raw.iterrows():
                row_upper = ' '.join(str(v).upper() for v in row.values if pd.notna(v))
                if 'MATERIAL' in row_upper and 'CURR' in row_upper:
                    return sheet_name
        except Exception:
            continue
    return xl.sheet_names[0]


def match_column(columns, patterns):
    """Exact pattern match — returns first column that contains any pattern."""
    for col in columns:
        col_upper = str(col).upper().replace('.', '').replace('_', ' ').strip()
        for pattern in patterns:
            if pattern in col_upper:
                return col
    return None


def find_near_matches(columns, patterns, threshold=2):
    """
    Find columns that are close but not exact matches.
    Used to give helpful suggestions when a required column is missing.
    Returns list of (column, pattern, distance) tuples.
    """
    near = []
    for col in columns:
        col_upper = str(col).upper().replace('.', '').replace('_', ' ').strip()
        for pattern in patterns:
            # Check if column contains most of the pattern (missing 1-2 chars)
            if len(pattern) > 4:
                for i in range(len(pattern) - 3):
                    fragment = pattern[i:i+4]
                    if fragment in col_upper:
                        near.append((col, pattern))
                        break
    return near


def diagnose_missing_columns(df, missing_fields):
    """
    When required columns are missing, provide detailed diagnostics:
    - Show all available columns
    - Suggest near-matches
    - Check if columns shifted (header row in wrong place)
    - Check if file has data at all
    """
    print(f"\n  {'─'*55}")
    print(f"  DIAGNOSIS — Could not find: {missing_fields}")
    print(f"  {'─'*55}")
    print(f"\n  Columns found in file ({len(df.columns)} total):")
    for i, col in enumerate(df.columns):
        print(f"    [{i+1:2d}] '{col}'")

    # Look for near-matches
    suggestions = {}
    for field in missing_fields:
        patterns = COL_PATTERNS[field]
        near = find_near_matches(df.columns, patterns)
        if near:
            suggestions[field] = near

    if suggestions:
        print(f"\n  Possible matches (columns with similar names):")
        for field, matches in suggestions.items():
            for col, pattern in matches[:2]:
                print(f"    '{field}' might be column '{col}' "
                      f"(looking for pattern '{pattern}')")

    # Check if data looks shifted — first column has numbers where material should be
    first_col_sample = df.iloc[:3, 0].tolist()
    if all(str(v).replace('.', '').replace(',', '').isdigit()
           for v in first_col_sample if pd.notna(v)):
        log_warn("First column appears to be numeric — header row may be misidentified. "
                 "Check if there are extra rows above the real header in the file.")

    # Check if file looks empty
    if len(df) < 5:
        log_warn(f"Only {len(df)} data rows found — file may be empty or filtered.")

    print(f"\n  Fix suggestions:")
    print(f"    1. Open the SAP export and check column headers are in row 1")
    print(f"       (or that no title/filter rows appear above the data)")
    print(f"    2. Verify the export includes: Material, Order Quantity, Net Value, Currency")
    print(f"    3. If SAP was reconfigured, update COL_PATTERNS in sap_cleaner.py")
    print(f"       to include the new column name")
    print(f"  {'─'*55}\n")


def is_subtotal_row(row, material_col):
    mat_val = str(row.get(material_col, '')).strip().lower()
    if any(ind in mat_val for ind in SUBTOTAL_INDICATORS):
        return True
    for val in row.values:
        if pd.notna(val) and any(ind in str(val).strip().lower()
                                  for ind in ['subtotal', 'grand total', '***']):
            return True
    return False


def normalize_currency(val):
    if pd.isna(val):
        return 'UNKNOWN'
    cleaned = str(val).strip().upper()
    return CURRENCY_NORMALIZATIONS.get(cleaned, cleaned)


def detect_number_format(df, col):
    samples = df[col].dropna().head(20).astype(str)
    european_votes = 0
    standard_votes = 0
    for val in samples:
        comma_pos  = val.rfind(',')
        period_pos = val.rfind('.')
        if comma_pos > period_pos > 0:
            european_votes += 1
        elif period_pos > comma_pos >= 0:
            standard_votes += 1
    if european_votes > standard_votes:
        log_warn(f"European number format detected in '{col}' "
                 f"({european_votes} samples) — adjusting parser")
        return True
    return False


def run(sap_path):
    print("=" * 60)
    print(f"SAP Data Cleaner  (pipeline v{PIPELINE_VERSION})")
    print("=" * 60)

    if VALIDATION_AVAILABLE:
        log_section("Validating inputs...")
        assert_file_exists(sap_path, 'SAP export')
        output_path = Path(sap_path).parent / 'sap_clean.xlsx'
        assert_output_dir_writable(str(output_path))

    print(f"\nReading: {sap_path}")

    try:
        xl = pd.ExcelFile(sap_path)
    except Exception as e:
        log_error(f"Could not open SAP file: {e}")
        print(f"\n  Fix: Make sure the file is not open in Excel and is a valid .xlsx file.")
        return

    # Find the right sheet
    sheet_name = find_data_sheet(xl)
    if len(xl.sheet_names) > 1:
        log_warn(f"Multiple sheets found {xl.sheet_names} — using '{sheet_name}'")

    # Find header row
    raw = xl.parse(sheet_name, header=None)

    header_row = find_header_row(raw, sheet_name)

    if header_row is None:
        log_error("Could not find header row in SAP export.")
        print(f"\n  Scanned first 30 rows looking for 'Material' and 'Currency' headers.")
        print(f"\n  Fix suggestions:")
        print(f"    1. Check the file is a SAP billing export, not a different report type")
        print(f"    2. Verify column headers are present and not all blank")
        print(f"    3. If headers are below row 30, move them up in the file")
        print(f"\n  First 5 rows of file for reference:")
        for i, row in raw.head(5).iterrows():
            vals = [str(v) for v in row.values if pd.notna(v)]
            print(f"    Row {i+1}: {vals}")
        return

    df = xl.parse(sheet_name, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    # Drop fully blank rows
    before_blank = len(df)
    df = df.dropna(how='all').reset_index(drop=True)
    blank_dropped = before_blank - len(df)
    if blank_dropped > 0:
        log_warn(f"Dropped {blank_dropped:,} blank rows from mid-export")

    print(f"  Raw rows loaded: {len(df):,}")
    print(f"  Columns found:   {list(df.columns)}")

    # Map required columns
    col_map = {}
    for field, patterns in COL_PATTERNS.items():
        col = match_column(df.columns, patterns)
        if col:
            col_map[field] = col

    missing = [f for f in COL_PATTERNS if f not in col_map]
    if missing:
        log_error(f"Could not find required column(s): {missing}")
        diagnose_missing_columns(df, missing)
        return

    # Warn on any required column that was a weak match
    for field, col in col_map.items():
        col_upper = str(col).upper().replace('.', '').replace('_', ' ').strip()
        best_pattern = COL_PATTERNS[field][0]
        if best_pattern not in col_upper:
            log_warn(f"'{field}' mapped to '{col}' via fallback pattern — "
                     f"verify this is the correct column")

    # Map context columns
    context_map = {}
    for field, patterns in CONTEXT_PATTERNS.items():
        col = match_column(df.columns, patterns)
        if col:
            context_map[field] = col

    print(f"\n  Column mapping:")
    for field, col in col_map.items():
        print(f"    {field:<15} -> '{col}'")
    for field, col in context_map.items():
        print(f"    {field:<15} -> '{col}' (context)")

    # Detect number format
    european_fmt = detect_number_format(df, col_map['net_value'])

    # Drop subtotal rows
    subtotal_mask  = df.apply(lambda r: is_subtotal_row(r, col_map['material']), axis=1)
    subtotal_count = subtotal_mask.sum()
    if subtotal_count > 0:
        log_warn(f"Dropped {subtotal_count:,} subtotal/total rows injected by SAP")
        df = df[~subtotal_mask].reset_index(drop=True)

    # Build clean DataFrame
    clean = pd.DataFrame()
    clean['material']  = df[col_map['material']].astype(str).str.strip()
    clean['quantity']  = df[col_map['quantity']].apply(
        lambda v: clean_number(v, european_fmt)
    )
    clean['quantity']  = pd.to_numeric(clean['quantity'], errors='coerce').fillna(0)
    clean['net_value'] = df[col_map['net_value']].apply(
        lambda v: clean_number(v, european_fmt)
    )
    clean['currency']  = df[col_map['currency']].apply(normalize_currency)

    for field, col in context_map.items():
        clean[field] = df[col].astype(str).str.strip()

    # Drop bad rows
    before = len(clean)
    clean  = clean[
        clean['material'].notna() &
        (~clean['material'].isin(['nan', 'none', 'NAN', 'NONE', '']))
    ]
    dropped_material = before - len(clean)

    after_material = len(clean)
    clean          = clean[clean['net_value'].notna()]
    dropped_net    = after_material - len(clean)

    # Warn on unexpected currency codes
    known      = set(CURRENCY_NORMALIZATIONS.values())
    unknown    = set(clean['currency'].unique()) - known
    if unknown:
        log_warn(f"Unknown currency codes: {unknown} — these rows will not match pricebook")

    # Warn if suspiciously few rows
    if len(clean) < 10:
        log_warn(f"Only {len(clean)} clean rows — verify the correct file was loaded")

    zero_qty = (clean['quantity'] == 0).sum()

    print(f"\n  Rows dropped:")
    print(f"    Missing material:   {dropped_material:,}")
    print(f"    Missing net value:  {dropped_net:,}")
    if subtotal_count > 0:
        print(f"    Subtotal rows:      {subtotal_count:,}")
    print(f"    Clean rows:         {len(clean):,}")
    print(f"    (qty = 0):          {zero_qty:,}  <- one-time fees, no tier lookup needed")

    print(f"\n  Summary:")
    print(f"    Unique materials:  {clean['material'].nunique():,}")
    print(f"    Currencies:        {', '.join(sorted(clean['currency'].unique()))}")
    print(f"    Net value range:   ${clean['net_value'].min():,.2f} - "
          f"${clean['net_value'].max():,.2f}")
    tiered = clean[clean['quantity'] > 0]
    if not tiered.empty:
        print(f"    Quantity range:    {int(tiered['quantity'].min()):,} - "
              f"{int(tiered['quantity'].max()):,}  (tiered rows only)")

    output_path = Path(sap_path).parent / 'sap_clean.xlsx'
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        clean.to_excel(writer, sheet_name='SAP', index=False)
        ws = writer.sheets['SAP']
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', start_color='1F4E79')

    print(f"\n  Output saved: {output_path}")
    print("=" * 60)


if __name__ == '__main__':
    run(SAP_FILE)
